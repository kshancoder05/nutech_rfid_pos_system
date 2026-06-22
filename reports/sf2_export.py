#!/usr/bin/env python3
"""
SF2 Exporter — fills the DepEd School Form 2 (Daily Attendance Report of Learners)
template from attendance data.

Data source (either):
  1. A DTR CSV with columns: Name, Sex, Date, Status   (Status = Present | Tardy | Absent)
  2. Live Airtable DTR table  (set AIRTABLE_PAT env var; uses --base/--students-dept)

Usage:
  python sf2_export.py --template SF2.xlsx --out SF2_filled.xlsx \
      --month 2026-06 --school "Lipa City NHS" --school-id 301234 \
      --grade "Grade 7" --section "Rizal" --csv attendance.csv

If no --csv and no AIRTABLE_PAT, a small built-in sample is used so you can see the layout.
"""
import argparse, calendar, csv, os, sys, datetime as dt
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

DAY_LETTER = {0:"M",1:"T",2:"W",3:"Th",4:"F",5:"Sat",6:"Sun"}
DATE_COL_START = 4           # column D
DATE_COL_END   = 28          # column AB
COL_ABSENT     = 29          # AC
COL_TARDY      = 30          # AD
ABSENT_MARK    = "x"
TARDY_MARK     = "T"

SAMPLE = [
    # Name, Sex, {date: status}
    ("Dela Cruz, Juan",  "Male",   {}),
    ("Santos, Pedro",    "Male",   {"absent":[3,17]}),
    ("Reyes, Mark",      "Male",   {"tardy":[10]}),
    ("Aquino, Maria",    "Female", {}),
    ("Bautista, Ana",    "Female", {"absent":[4]}),
    ("Garcia, Liza",     "Female", {"tardy":[2,18],"absent":[24]}),
]

def school_days(year, month):
    """Weekday dates (Mon-Fri) of the month, as day-of-month ints."""
    n = calendar.monthrange(year, month)[1]
    days = []
    for d in range(1, n+1):
        wd = dt.date(year, month, d).weekday()
        if wd < 5:  # Mon-Fri
            days.append((d, DAY_LETTER[wd]))
    return days

def unmerge_region(ws, r1, r2, c1, c2):
    """Unmerge any merged ranges intersecting the rows/cols box so we can write cells."""
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= r1 and rng.max_row <= r2 and rng.min_col >= c1 and rng.max_col <= c2:
            ws.unmerge_cells(str(rng))

def set_cell(ws, row, col, value):
    """Write a value, redirecting to the top-left cell if the target is inside a merge."""
    c = ws.cell(row=row, column=col)
    if isinstance(c, MergedCell):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                ws.cell(row=rng.min_row, column=rng.min_col).value = value
                return
        return
    c.value = value

def find_block_rows(ws):
    """Locate MALE / FEMALE total rows by scanning column A, return learner row ranges."""
    male_total = female_total = None
    for r in range(12, ws.max_row+1):
        v = str(ws.cell(row=r, column=1).value or "").upper()
        if "MALE" in v and "TOTAL" in v and "FEMALE" not in v and male_total is None:
            male_total = r
        elif "FEMALE" in v and "TOTAL" in v and female_total is None:
            female_total = r
    male_rows   = list(range(14, male_total)) if male_total else []
    female_rows = list(range(male_total+1, female_total)) if (male_total and female_total) else []
    return male_rows, female_rows

def load_csv(path):
    people = defaultdict(lambda: {"sex":"Male","marks":{}})
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            name = (row.get("Name") or "").strip()
            if not name: continue
            people[name]["sex"] = (row.get("Sex") or "Male").strip().title()
            d = (row.get("Date") or "").strip()
            st = (row.get("Status") or "Present").strip().title()
            if d:
                try: day = dt.date.fromisoformat(d).day
                except ValueError: continue
                people[name]["marks"][day] = st
    return [(n, v["sex"], v["marks"]) for n, v in people.items()]

def load_airtable(base, month_start, month_end):
    import urllib.request, json
    pat = os.environ["AIRTABLE_PAT"]
    def fetch(table, formula=None):
        url = f"https://api.airtable.com/v0/{base}/{table}?pageSize=100"
        if formula: url += "&filterByFormula=" + urllib.parse.quote(formula)
        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {pat}"})
        return json.load(urllib.request.urlopen(req)).get("records", [])
    people = defaultdict(lambda: {"sex":"Male","marks":{}})
    for rec in fetch("DTR"):
        f = rec["fields"]
        if (f.get("Department") or "") != "Student": continue
        d = f.get("Date","")
        if not (month_start <= d <= month_end): continue
        name = f.get("Name","")
        if not name: continue
        people[name]["marks"][dt.date.fromisoformat(d).day] = f.get("Status","Present")
    return [(n, v["sex"], v["marks"]) for n, v in people.items()]

def normalize(people, days):
    """Turn each person's marks into absent/tardy day-sets across school days."""
    daynums = [d for d,_ in days]
    out = []
    for name, sex, marks in people:
        absent, tardy = set(), set()
        if "absent" in marks or "tardy" in marks:          # SAMPLE shape
            absent = set(marks.get("absent", []))
            tardy  = set(marks.get("tardy", []))
        else:                                              # {day: status} shape
            for d in daynums:
                st = str(marks.get(d, "")).title()
                if st == "Absent": absent.add(d)
                elif st == "Tardy": tardy.add(d)
        out.append({"name":name, "sex":sex, "absent":absent, "tardy":tardy})
    return out

def fill(template, out, year, month, school, school_id, grade, section, people):
    wb = load_workbook(template)
    ws = wb.active
    days = school_days(year, month)

    # header (write into the merged input cells, not the labels)
    set_cell(ws, 6, 3,  school_id)              # School ID input  (C6:E6)
    set_cell(ws, 6, 11, f"{year}-{year+1}")     # School Year      (K6:O6)
    set_cell(ws, 6, 24, dt.date(year,month,1).strftime("%B %Y"))  # Month (X6:AC6)
    set_cell(ws, 8, 3,  school)                 # Name of School
    set_cell(ws, 8, 24, grade)                  # Grade Level input (X8:Y8)
    set_cell(ws, 8, 29, section)                # Section input    (AC8:AH8)

    # date header row 10 (numbers) + row 11 (day letters)
    unmerge_region(ws, 10, 11, DATE_COL_START, DATE_COL_END)
    for i,(d,letter) in enumerate(days):
        col = DATE_COL_START + i
        if col > DATE_COL_END: break
        ws.cell(row=10, column=col).value = d
        ws.cell(row=11, column=col).value = letter

    male_rows, female_rows = find_block_rows(ws)
    males   = [p for p in people if p["sex"].lower().startswith("m")]
    females = [p for p in people if p["sex"].lower().startswith("f")]

    def write_block(plist, rows):
        for p, r in zip(plist, rows):
            ws.cell(row=r, column=1).value = p["name"]
            for i,(d,_) in enumerate(days):
                col = DATE_COL_START + i
                if col > DATE_COL_END: break
                if d in p["absent"]:   ws.cell(row=r, column=col).value = ABSENT_MARK
                elif d in p["tardy"]:  ws.cell(row=r, column=col).value = TARDY_MARK
            ws.cell(row=r, column=COL_ABSENT).value = len(p["absent"]) or None
            ws.cell(row=r, column=COL_TARDY).value  = len(p["tardy"]) or None

    write_block(males, male_rows)
    write_block(females, female_rows)
    wb.save(out)
    return len(males), len(females), len(days)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", default="SF2_filled.xlsx")
    ap.add_argument("--month", required=True, help="YYYY-MM")
    ap.add_argument("--school", default="")
    ap.add_argument("--school-id", default="")
    ap.add_argument("--grade", default="")
    ap.add_argument("--section", default="")
    ap.add_argument("--csv", default=None)
    ap.add_argument("--base", default="apprpYxg7leO7JXKJ")
    a = ap.parse_args()
    year, month = int(a.month[:4]), int(a.month[5:7])

    if a.csv:
        raw = load_csv(a.csv)
    elif os.environ.get("AIRTABLE_PAT"):
        last = calendar.monthrange(year, month)[1]
        raw = load_airtable(a.base, f"{a.month}-01", f"{a.month}-{last:02d}")
    else:
        raw = SAMPLE
        print("No --csv or AIRTABLE_PAT: using built-in sample data.")

    days = school_days(year, month)
    people = normalize(raw, days)
    m,f,nd = fill(a.template, a.out, year, month, a.school, a.school_id, a.grade, a.section, people)
    print(f"Wrote {a.out}: {m} male + {f} female learners across {nd} school days.")

if __name__ == "__main__":
    main()
